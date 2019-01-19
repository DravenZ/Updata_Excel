# -*- coding: utf-8 -*-
import os
import uuid
import platform
from flask import Flask, request, redirect, url_for, render_template, send_from_directory, abort, make_response
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

if platform.system() == "Windows":
    slash = '\\'
else:
    platform.system() == "Linux"
    slash = '/'
UPLOAD_FOLDER = 'upload'
ALLOW_EXTENSIONS = set(['xlsx', 'xls'])
app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
# 判断文件夹是否存在，如果不存在则创建
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
else:
    pass


# 判断文件后缀是否在列表中
def allowed_file(filename):
    return '.' in filename and \
            filename.rsplit('.', 1)[1] in ALLOW_EXTENSIONS


@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        # 获取post过来的文件名称，从name=file参数中获取
        flag = 0
        if request.files.get('file1') is not None:
            file = request.files.get('file1')
            flag = 1
        else:
            file = request.files.get('file2')
        if file and allowed_file(file.filename):

            filename = '标记' + file.filename
            # file_name = str(uuid.uuid4()) + '.' + filename.rsplit('.', 1)[1]
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            base_path = os.getcwd().split(slash)[-1]
            file_path = base_path + slash + app.config['UPLOAD_FOLDER'] + slash + filename
            print(file_path)

            fill = PatternFill("solid", fgColor="FFC000")
            wb = load_workbook(app.config['UPLOAD_FOLDER'] + slash + filename)
            wnames = wb.get_sheet_names()
            print(wnames)
            for wname in wnames:
                sheet = wb.get_sheet_by_name(wname)
                # sheet = wb.active
                maxrow = sheet.max_row
                maxcol = sheet.max_column
                for i in range(maxrow):
                    for j in range(maxcol):
                        cell = sheet.cell(row=i + 1, column=j + 1)
                        if j == 0 and cell.value is None:
                            cell.fill = fill
                        if flag == 0:
                            if j == 1 and cell.value is None:
                                cell.fill = fill
                        if len(str(cell.value)) > 200:
                            cell.fill = fill
                        # 这个是植物资料库的喜阴喜阳为空判断
                        if flag == 1:
                            if j == 7 and (cell.value is None or cell.value not in (0, 1)):
                                cell.fill = fill

            # wb.save(filename)
            wb.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            return render_template('base.html', filename=filename, flag=flag)
        else:
            return redirect(url_for('upload_file'))

    return render_template('base.html')


@app.route('/download/<filename>', methods=['GET'])
def download(filename):
    if request.method == "GET":
        if os.path.isfile(os.path.join('upload', filename)):
            response = make_response(send_from_directory('upload', filename, as_attachment=True))
            response.headers["Content-Disposition"] = "attachment; filename={}".format(
                filename.encode().decode('latin-1'))
            return response
        abort(404)


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
